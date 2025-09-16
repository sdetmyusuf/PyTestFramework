import openpyxl
from pygments.lexer import words


def getRowCount (sheetPath, sheetName):
    workbook = openpyxl.load_workbook(sheetPath)
    sheet = workbook[sheetName]
    return sheet.max_row

def getColumnCount (sheetPath, sheetName):
    workbook = openpyxl.load_workbook(sheetPath)
    sheet = workbook[sheetName]
    return sheet.max_column


def getCellData (sheetPath, sheetName):
    workbook = openpyxl.load_workbook(sheetPath)
    sheet = workbook[sheetName]
    return sheet.cell()

def setCellData (sheetPath, sheetName, row_number, col_number, data):
    workbook = openpyxl.load_workbook(sheetPath)
    sheet = workbook[sheetName]
    sheet.cell(row=row_number, column=col_number).value = data
    workbook.save(sheetPath)

def getDataFromExcell (sheetPath, sheetName):
    final_list = []
    workbook = openpyxl.load_workbook(sheetPath)
    sheet = workbook[sheetName]
    total_rows = sheet.max_row
    total_cols = sheet.max_column
    print("total rows", total_rows)
    print("total columns", total_cols)
    for r in range (2, total_rows+1):
        row_list = []
        for c in range(1, total_cols+1):
            row_list.append(sheet.cell(row=r, column=c).value)
        final_list.append(row_list)

    return final_list


lst = getDataFromExcell("C://Users//Mohd Yusuf//PycharmProjects//PYTestFramework//ExcelFiles//LoginData.xlsx", "LoginData")
print(lst)
